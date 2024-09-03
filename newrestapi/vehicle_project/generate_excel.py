import requests
import pandas as pd
from datetime import datetime, timedelta
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

LOGIN_URL = "https://api.baubuddy.de/index.php/login"
DATA_URL = "https://api.baubuddy.de/dev/index.php/v1/vehicles/select/active"
API_KEY = "QVBJX0V4cGxvcmVyOjEyMzQ1NmlzQUxhbWVQYXNz"
USERNAME = "365"
PASSWORD = "1"

def get_access_token():
    headers = {
        "Authorization": f"Basic {API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "username": USERNAME,
        "password": PASSWORD
    }
    response = requests.post(LOGIN_URL, json=payload, headers=headers)
    response_data = response.json()
    return response_data['oauth']['access_token']

def fetch_data(access_token):
    headers = {
        "Authorization": f"Bearer {access_token}"
    }
    response = requests.get(DATA_URL, headers=headers)
    return response.json()

def save_to_excel(data, file_name, keys, colored):
    df = pd.DataFrame(data)
    
    df.columns = df.columns.str.strip()
    
    df['rnr'] = pd.to_numeric(df['rnr'], errors='coerce')

    columns_to_include = ['rnr'] + keys
    df = df[columns_to_include]

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    ws.append(columns_to_include)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center')

    green_fill = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="B30000", end_color="B30000", fill_type="solid")

    for idx, row in df.iterrows():
        hu_date = pd.to_datetime(row.get('hu', ''), errors='coerce') if 'hu' in row else None

        if colored and hu_date:
            now = pd.Timestamp.now()
            if hu_date > now - timedelta(days=90): 
                fill = green_fill
            elif hu_date > now - timedelta(days=365):  
                fill = orange_fill
            else:
                fill = red_fill
        else:
            fill = None

        ws.append(row[columns_to_include].to_list())

        if fill:
            for col in range(1, len(columns_to_include) + 1):
                ws.cell(row=idx + 2, column=col).fill = fill

    wb.save(file_name)
    print(f"Excel dosyası kaydedildi: {file_name}")

def main():
    parser = argparse.ArgumentParser(description="Process vehicle data.")
    parser.add_argument('-k', '--keys', nargs='+', help="Additional columns to include.")  
    parser.add_argument('-c', '--colored', action='store_true', help="Color the Excel output.")
    args = parser.parse_args()

    access_token = get_access_token()
    
    data = fetch_data(access_token)
    
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f'vehicles_{today}.xlsx'
    
    save_to_excel(data, filename, args.keys, args.colored)

if __name__ == "__main__":
    main()
